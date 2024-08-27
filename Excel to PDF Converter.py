# -*- coding: utf-8 -*-
import os
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinterdnd2 import TkinterDnD, DND_FILES  # TkinterDnD2支持拖放
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from PIL import Image, ImageTk

def browse_folder():
    folder_selected = filedialog.askdirectory()
    if folder_selected:
        folder_path.set(folder_selected)

def browse_output_folder():
    output_selected = filedialog.askdirectory()
    if output_selected:
        output_folder_path.set(output_selected)

def split_paragraph_to_fit_cell(paragraph, style, max_height):
    lines = paragraph.split('\n')
    chunks = []
    current_chunk = []
    current_height = 0
    
    for line in lines:
        p = Paragraph(line, style)
        current_chunk.append(p)
        current_height += p.wrap(0, 0)[1]  # get height
        if current_height >= max_height:
            chunks.append(current_chunk)
            current_chunk = []
            current_height = 0
    
    if current_chunk:
        chunks.append(current_chunk)
    
    return chunks

def convert_excel_to_pdf(input_folder, output_folder):
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    for filename in os.listdir(input_folder):
        if filename.endswith(".xlsx") and not filename.startswith('~$'):  # Skip temporary files
            input_path = os.path.join(input_folder, filename)
            output_path = os.path.join(output_folder, filename.replace(".xlsx", ".pdf"))

            print(f"Processing file: {input_path}")

            try:
                workbook = load_workbook(input_path)
                sheet = workbook.active

                # Prepare data for PDF
                data = []
                col_widths = [1.5 * inch, 1.0 * inch, 2.5 * inch, 1.5 * inch]  # 手动设置列宽
                styles = getSampleStyleSheet()
                default_style = ParagraphStyle(
                    name="TableCell",
                    fontName="Helvetica",
                    fontSize=7,  # Default font size
                    leading=8,
                    alignment=1  # 居中对齐
                )

                special_style = ParagraphStyle(
                    name="SpecialCell",
                    fontName="Helvetica",
                    fontSize=12,  # Larger font size for special cases
                    leading=14,
                    alignment=1
                )

                max_cell_height = 300  # Define max height for each cell
                elements = []

                for row in sheet.iter_rows(values_only=True):
                    row_data = []
                    for i, item in enumerate(row):
                        if item is not None:
                            content = str(item)
                            p = Paragraph(content, default_style)
                            if p.wrap(col_widths[i], max_cell_height)[1] > max_cell_height:
                                elements.append(PageBreak())
                                elements.append(Paragraph(content, special_style))
                                elements.append(PageBreak())
                            else:
                                row_data.append(p)
                        else:
                            row_data.append("")
                    data.append(row_data)

                if not data:
                    print(f"No data found in {filename}, skipping.")
                    continue

                # Generate table for the main content
                table = Table(data, colWidths=col_widths)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 7),  # Smaller font size
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),  # 内容顶部对齐
                ]))

                elements.insert(0, table)  # Place table before any new pages
                pdf = SimpleDocTemplate(output_path, pagesize=landscape(letter))
                pdf.build(elements)

                print(f"Successfully converted {filename} to PDF.")

            except Exception as e:
                print(f"Failed to process {filename}: {str(e)}")

    print("Excel files conversion to PDF process completed.")

def start_conversion():
    folder = folder_path.get()
    output_folder = output_folder_path.get()
    
    if not folder:
        messagebox.showerror("Error", "Please select a folder!")
        return
    
    if not output_folder:
        messagebox.showerror("Error", "Please select an output folder!")
        return
    
    convert_excel_to_pdf(folder, output_folder)
    messagebox.showinfo("Info", f"Conversion completed. PDF files saved in: {output_folder}")

def drop(event):
    # 处理拖放的文件夹路径
    folder_selected = event.data.strip('{}')  # 去掉路径中的大括号
    if os.path.isdir(folder_selected):
        folder_path.set(folder_selected)

def create_gui():
    global folder_path, output_folder_path
    root = TkinterDnD.Tk()  # 使用TkinterDnD2扩展以支持拖放
    root.title("Excel to PDF Converter")
    
    # 设置窗口图标
    logo_path = "C:/Users/admin/Desktop/logo.png"  # 使用你上传的logo图片
    if os.path.exists(logo_path):
        img = Image.open(logo_path)
        img = img.resize((32, 32))  # 图标一般为32x32像素
        logo = ImageTk.PhotoImage(img)
        root.iconphoto(False, logo)  # 设置窗口图标

    folder_path = tk.StringVar()
    output_folder_path = tk.StringVar()

    tk.Label(root, text="Drag and drop a folder with Excel files or browse:").pack(pady=10)

    entry = tk.Entry(root, textvariable=folder_path, width=50)
    entry.pack(pady=5)

    browse_button = tk.Button(root, text="Browse Input Folder", command=browse_folder)
    browse_button.pack(pady=5)

    tk.Label(root, text="Select the output folder:").pack(pady=10)

    output_entry = tk.Entry(root, textvariable=output_folder_path, width=50)
    output_entry.pack(pady=5)

    output_browse_button = tk.Button(root, text="Browse Output Folder", command=browse_output_folder)
    output_browse_button.pack(pady=5)

    convert_button = tk.Button(root, text="Start Conversion", command=start_conversion)
    convert_button.pack(pady=20)

    # 支持拖放文件夹
    root.drop_target_register(DND_FILES)
    root.dnd_bind('<<Drop>>', drop)

    root.mainloop()

if __name__ == "__main__":
    create_gui()

